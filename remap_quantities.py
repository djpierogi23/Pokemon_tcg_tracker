#!/usr/bin/env python3
"""
Re-extract collection data with proper quantity mapping.
Columns are always: Set#(0), Name(1), Type(2), Qty1(3), Qty2(4), Qty3(5), Qty4(6), Note(7), Have?(8), Stock(9), Playable(10), Rarity(11)
"""
import openpyxl
import json
import re

XLSX = "/Users/danodomirok/.gemini/antigravity/scratch/Pokémon TCG Spreadsheet - Collector's Edition (Improved).xlsx"

GENERATIONS = [
    {"id": "gen1", "name": "Gen 1 (Red and Blue)", "color": "#E74C3C",
     "sets": ["Wizards Black Star Promos", "Base Set (1st Edition)", "Base Set (Shadowless)",
              "Base Set", "Base Set (Japanese)", "Jungle", "Fossil", "Base Set 2",
              "Team Rocket", "Gym Heroes", "Gym Challenge",
              "Challenge from the Darkness (Ja", "Legendary Collection",
              "Vending Series 1 (J)", "Vending Series 2 (J)", "Vending Series 3 (J)"]},
    {"id": "gen2", "name": "Gen 2 (Gold and Silver)", "color": "#F1C40F",
     "sets": ["Neo Genesis", "Gold, Silver, to a New World (J",
              "Neo Discovery", "Southern Islands", "Neo Revelation", "Neo Destiny",
              "Expedition Base Set", "Aquapolis", "Skyridge", "VS (J)"]},
    {"id": "gen3", "name": "Gen 3 (Ruby and Sapphire)", "color": "#2ECC71",
     "sets": ["EX Ruby & Sapphire", "EX Sandstorm", "EX Dragon", "Magma vs Aqua",
              "EX Hidden Legends", "EX FireRed & LeafGreen", "EX Team Rocket Returns",
              "EX Deoxys", "EX Emerald", "EX Unseen Forces", "EX Delta Species",
              "EX Legend Maker", "EX Holon Phantoms", "EX Crystal Guardians",
              "EX Dragon Frontiers", "EX Power Keepers"]},
    {"id": "gen4", "name": "Gen 4 (Diamond and Pearl)", "color": "#3498DB",
     "sets": ["Diamond & Pearl", "Mysterious Treasures", "Secret Wonders",
              "Great Encounters", "Majestic Dawn", "Legends Awakened",
              "Stormfront", "Platinum", "Rising Rivals", "Supreme Victors",
              "Arceus", "HGSS", "Unleashed", "Undaunted", "Triumphant", "Call of Legends"]},
    {"id": "gen5", "name": "Gen 5 (Black and White)", "color": "#9B59B6",
     "sets": ["Black & White", "Emerging Powers", "Noble Victories",
              "Next Destinies", "Dark Explorers", "Dragons Exalted",
              "Dragon Vault", "Boundaries Crossed", "Plasma Storm",
              "Plasma Freeze", "Plasma Blast", "Legendary Treasures",
              "Black Bolt", "White Flare"]},
    {"id": "gen6", "name": "Gen 6 (X and Y)", "color": "#E67E22",
     "sets": ["XY", "Flashfire", "Furious Fists", "Phantom Forces",
              "Primal Clash", "Roaring Skies", "Ancient Origins",
              "BREAKthrough", "BREAKpoint", "Fates Collide",
              "Steam Siege", "Evolutions", "Generations"]},
    {"id": "gen7", "name": "Gen 7 (Sun and Moon)", "color": "#1ABC9C",
     "sets": ["Sun & Moon", "Guardians Rising", "Burning Shadows",
              "Shining Legends", "Crimson Invasion", "Ultra Prism",
              "Forbidden Light", "Celestial Storm", "Dragon Majesty",
              "Lost Thunder", "Team Up", "Detective Pikachu",
              "Unbroken Bonds", "Unified Minds", "Hidden Fates", "Cosmic Eclipse"]},
    {"id": "gen8", "name": "Gen 8 (Sword and Shield)", "color": "#1ABC9C",
     "sets": ["Sword & Shield", "Rebel Clash", "Darkness Ablaze",
              "Champion's Path", "Vivid Voltage", "Shining Fates",
              "Battle Styles", "Chilling Reign", "Evolving Skies",
              "Celebrations", "Fusion Strike", "Brilliant Stars",
              "Astral Radiance", "Pokemon GO", "Lost Origin",
              "Silver Tempest", "Crown Zenith"]},
    {"id": "gen9", "name": "Gen 9 (Scarlet and Violet)", "color": "#9B59B6",
     "sets": ["Scarlet & Violet", "Paldea Evolved", "Obsidian Flames",
              "151", "Paradox Rift", "Paldean Fates", "Temporal Forces",
              "Twilight Masquerade", "Shrouded Fable", "Stellar Crown",
              "Surging Sparks", "Prismatic Evolutions", "Journey Together"]},
    {"id": "gen10", "name": "Gen 10+", "color": "#16A085", "sets": []},
]

# Sets with special mapping rules
FIRST_EDITION_SETS = {"Base Set (1st Edition)"}
SHADOWLESS_SETS = {"Base Set (Shadowless)"}

# Sets with Quantity-1st Edition in col[6] (pattern 2)
FIRST_ED_COL6_SETS = {"Jungle", "Fossil", "Team Rocket", "Gym Heroes", "Gym Challenge",
                       "Neo Genesis", "Neo Discovery", "Neo Revelation", "Neo Destiny"}

def to_int(val):
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0

def slugify(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

def map_quantities(sheet_name, qty_headers, q3, q4, q5, q6):
    """Map the 4 quantity columns (indices 3-6) to API variant types."""
    quantities = {"normal": 0, "holofoil": 0, "reverseHolofoil": 0, "firstEdition": 0, "unlimited": 0}
    
    if sheet_name in FIRST_EDITION_SETS:
        # Pattern: [Qty-1st Edition, Qty-Normal, Qty-Holo, Qty-Promo]
        quantities["firstEdition"] = q3
        quantities["normal"] = q4
        quantities["holofoil"] = q5
        
    elif sheet_name in SHADOWLESS_SETS:
        # Pattern: [Qty-1st Edition, Qty-Normal, Qty-Holo, Qty-Promo]
        # Shadowless are unlimited variants
        quantities["firstEdition"] = q3
        quantities["unlimited"] = q4
        quantities["holofoil"] = q5
        
    elif sheet_name in FIRST_ED_COL6_SETS:
        # Pattern: [Qty-Normal, Qty-Reverse, Qty-Holo, Qty-1st Edition]
        # Normal cards are unlimited (non-1st-ed)
        quantities["unlimited"] = q3
        quantities["reverseHolofoil"] = q4
        quantities["holofoil"] = q5
        quantities["firstEdition"] = q6
        
    else:
        # Standard pattern (146 sets): [Qty-Normal, Qty-Reverse, Qty-Holo, Qty-Promo]
        quantities["normal"] = q3
        quantities["reverseHolofoil"] = q4
        quantities["holofoil"] = q5
        # Promo goes to unlimited for standard sets
        quantities["unlimited"] = q6
    
    return quantities

def extract():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    skip = {'Overview', 'Template'}
    all_sheets = [s for s in wb.sheetnames if s not in skip]
    
    # Assign sheets to generations
    assigned = set()
    for gen in GENERATIONS:
        gen["_matched"] = []
    
    for sheet_name in all_sheets:
        found = False
        for gen in GENERATIONS:
            if sheet_name in gen["sets"]:
                gen["_matched"].append(sheet_name)
                assigned.add(sheet_name)
                found = True
                break
        if not found:
            GENERATIONS[-1]["_matched"].append(sheet_name)
    
    output = {"generations": []}
    total_cards = 0
    total_with_qty = 0
    
    for gen in GENERATIONS:
        gen_data = {"id": gen["id"], "name": gen["name"], "color": gen["color"], "sets": []}
        
        for sheet_name in gen["_matched"]:
            ws = wb[sheet_name]
            
            set_name = ws.cell(row=1, column=1).value or sheet_name
            release_date = ws.cell(row=2, column=2).value or ""
            if hasattr(release_date, 'strftime'):
                release_date = release_date.strftime('%Y-%m-%d')
            else:
                release_date = str(release_date) if release_date else ""
            
            # Read quantity header names for reference
            qty_headers = []
            for col in range(4, 8):  # D, E, F, G (1-indexed: 4,5,6,7)
                val = ws.cell(row=4, column=col).value
                qty_headers.append(str(val).strip() if val else "")
            
            cards = []
            for row in ws.iter_rows(min_row=5, max_col=12, values_only=True):
                if not row or not row[0]:
                    continue
                
                card_number = str(row[0]) if row[0] else ""
                card_name = str(row[1]).strip() if row[1] else ""
                if not card_name:
                    continue
                
                card_type = str(row[2]).strip() if row[2] else ""
                
                # Quantities at indices 3,4,5,6
                q3 = to_int(row[3] if len(row) > 3 else None)
                q4 = to_int(row[4] if len(row) > 4 else None)
                q5 = to_int(row[5] if len(row) > 5 else None)
                q6 = to_int(row[6] if len(row) > 6 else None)
                
                quantities = map_quantities(sheet_name, qty_headers, q3, q4, q5, q6)
                
                # Note at index 7
                note = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                note = note.replace('\n', ' ').replace('\r', '')
                
                # Have? at index 8
                have_val = row[8] if len(row) > 8 else None
                if have_val:
                    hv = str(have_val).strip().upper()
                    status = "HAVE" if hv in ('Y', 'YES', 'TRUE', '1', 'HAVE') else "NEED"
                else:
                    status = "NEED"
                
                # Stock at index 9
                stock = to_int(row[9] if len(row) > 9 else None)
                
                # Rarity at index 11
                rarity = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                
                has_qty = any(v > 0 for v in quantities.values())
                if has_qty:
                    total_with_qty += 1
                elif status == "HAVE":
                    # Card is owned but no specific variant — assume normal
                    quantities["normal"] = max(stock, 1)
                    total_with_qty += 1
                
                cards.append({
                    "number": card_number,
                    "name": card_name,
                    "type": card_type,
                    "status": status,
                    "stock": stock,
                    "quantities": quantities,
                    "rarity": rarity,
                    "note": note,
                })
            
            total_cards += len(cards)
            gen_data["sets"].append({
                "id": slugify(sheet_name),
                "name": set_name if isinstance(set_name, str) else sheet_name,
                "sheetName": sheet_name,
                "releaseDate": release_date,
                "cards": cards,
            })
        
        if gen_data["sets"]:
            output["generations"].append(gen_data)
    
    wb.close()
    
    # Write files
    json_path = "/Users/danodomirok/.gemini/antigravity/scratch/Pokemon_tcg_tracker/data/collection.json"
    js_path = "/Users/danodomirok/.gemini/antigravity/scratch/Pokemon_tcg_tracker/data/collection.js"
    
    with open(json_path, 'w') as f:
        json.dump(output, f, ensure_ascii=False)
    with open(js_path, 'w') as f:
        f.write("var COLLECTION_DATA = ")
        json.dump(output, f, ensure_ascii=False)
        f.write(";\n")
    
    print(f"Extracted {total_cards} cards across {sum(len(g['sets']) for g in output['generations'])} sets")
    print(f"Cards with at least one quantity > 0: {total_with_qty}")
    
    # Verify a few sets
    for gen in output["generations"][:3]:
        for s in gen["sets"][:4]:
            qcards = [c for c in s["cards"] if any(v > 0 for v in c["quantities"].values())]
            if qcards:
                print(f"\n{s['name']}:")
                for c in qcards[:5]:
                    q = c["quantities"]
                    parts = [f"{k}={v}" for k, v in q.items() if v > 0]
                    print(f"  {c['name']}: {', '.join(parts)}")

if __name__ == "__main__":
    extract()
