#!/usr/bin/env python3
"""Analyze Excel sheet columns to understand quantity field mapping."""
import openpyxl
import json

wb = openpyxl.load_workbook(
    "/Users/danodomirok/.gemini/antigravity/scratch/Pokémon TCG Spreadsheet - Collector's Edition (Improved).xlsx",
    data_only=True, read_only=True
)

skip = {'Overview', 'Template'}
results = {}

for sheet_name in wb.sheetnames:
    if sheet_name in skip:
        continue
    ws = wb[sheet_name]
    # Read header row (row 4)
    headers = []
    for col in range(1, 15):
        cell = ws.cell(row=4, column=col)
        val = cell.value
        if val:
            headers.append(str(val).strip())
        else:
            break
    
    # Find quantity-related columns
    qty_cols = [h for h in headers if 'qty' in h.lower() or 'quantity' in h.lower() or 'quant' in h.lower()]
    results[sheet_name] = qty_cols

wb.close()

# Print grouped by column pattern
from collections import defaultdict
patterns = defaultdict(list)
for sheet, cols in results.items():
    key = tuple(cols)
    patterns[key].append(sheet)

for pattern, sheets in sorted(patterns.items(), key=lambda x: -len(x[1])):
    print(f"\n=== Pattern: {list(pattern)} ({len(sheets)} sets) ===")
    for s in sheets[:10]:
        print(f"  - {s}")
    if len(sheets) > 10:
        print(f"  ... and {len(sheets)-10} more")
