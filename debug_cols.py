#!/usr/bin/env python3
import openpyxl
wb = openpyxl.load_workbook("/Users/danodomirok/.gemini/antigravity/scratch/Pokémon TCG Spreadsheet - Collector's Edition (Improved).xlsx", data_only=True, read_only=True)

for sn in ['Jungle', 'Base Set (1st Edition)', 'Legendary Collection', 'Base Set']:
    ws = wb[sn]
    headers = []
    for col in range(1, 15):
        val = ws.cell(row=4, column=col).value
        headers.append(str(val).strip() if val else '')
    print(f'\n{sn} headers:')
    for i, h in enumerate(headers):
        if h:
            print(f'  col[{i}] = {h}')
    
    row5 = [ws.cell(row=5, column=c).value for c in range(1, 15)]
    print(f'  Row 5 data: {row5}')
wb.close()
